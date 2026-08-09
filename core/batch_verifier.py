"""Session-only batch verification support for crawled news files."""

from __future__ import annotations

import csv
import io
import json
import re
from collections import Counter, defaultdict
from collections.abc import Callable, Iterable
from dataclasses import dataclass
from datetime import date
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from core.claim_splitter import split_complex_claim
from schemas.verdict import VerdictSchema

_REQUIRED_COLUMNS = {"article_id", "published_at", "body"}
_NUMERIC = re.compile(r"\d")
_SENTENCES = re.compile(r"(?<=[.!?])\s+|\n+")


@dataclass(frozen=True, slots=True)
class BatchArticle:
    article_id: str
    published_at: date
    body: str
    title: str | None = None
    source_url: str | None = None


@dataclass(frozen=True, slots=True)
class BatchClaimResult:
    article_id: str
    published_at: date
    source_sentence: str
    claim_id: str
    verdict: str
    route_status: str
    reason_code: str
    claim_value: float | None
    calculated_value: float | None
    evidence_value: float | None
    kosis_table_id: str | None
    evidence_key: str | None
    source_url: str | None = None


@dataclass(frozen=True, slots=True)
class BatchArticleSummary:
    article_id: str
    published_at: date
    claim_count: int
    match_count: int
    mismatch_count: int
    review_count: int
    article_status: str
    source_url: str | None = None


@dataclass(frozen=True, slots=True)
class BatchVerificationResult:
    claim_rows: list[BatchClaimResult]
    article_rows: list[BatchArticleSummary]


Verifier = Callable[[str, date], VerdictSchema]


def load_articles(filename: str, content: bytes, *, default_published_at: date | None = None) -> list[BatchArticle]:
    """Decode an uploaded CSV, JSON, or XLSX in memory and normalize known crawler columns."""
    suffix = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    rows = [_canonicalize_row(row) for row in _read_rows(suffix, content)]
    for index, row in enumerate(rows, start=1):
        row.setdefault("article_id", f"row-{index:06d}")
        if "body" not in row and _optional_text(row.get("title")):
            row["body"] = row["title"]
    if not rows or not all(_optional_text(row.get("body")) for row in rows):
        raise ValueError("BATCH_REQUIRED_COLUMNS")
    if any(_optional_text(row.get("published_at")) is None for row in rows) and default_published_at is None:
        raise ValueError("BATCH_ARTICLE_DATE_REQUIRED")
    articles: list[BatchArticle] = []
    for index, row in enumerate(rows, start=2):
        try:
            article_id = _required_text(row, "article_id")
            published_value = _optional_text(row.get("published_at"))
            published_at = date.fromisoformat(published_value) if published_value else default_published_at
            if published_at is None:
                raise ValueError("published_at")
            body = _required_text(row, "body")
        except (TypeError, ValueError) as error:
            raise ValueError(f"BATCH_INVALID_ROW:{index}") from error
        articles.append(
            BatchArticle(
                article_id=article_id,
                published_at=published_at,
                body=body,
                title=_optional_text(row.get("title")),
                source_url=_optional_text(row.get("source_url")),
            )
        )
    return articles

def verify_articles(articles: Iterable[BatchArticle], verifier: Verifier) -> BatchVerificationResult:
    """Verify each numeric claim independently; a failed claim never aborts the batch."""
    claim_rows: list[BatchClaimResult] = []
    materialized = list(articles)
    for article in materialized:
        for sentence in _numeric_claims(article.body):
            try:
                verdict = verifier(sentence, article.published_at)
                claim_rows.append(_claim_row(article, sentence, verdict))
            except Exception:
                claim_rows.append(_hold_row(article, sentence))
    return BatchVerificationResult(claim_rows=claim_rows, article_rows=_summaries(materialized, claim_rows))


def export_batch_xlsx(result: BatchVerificationResult) -> bytes:
    """Build a download-ready review workbook entirely in memory."""
    workbook = Workbook()
    claims = workbook.active
    claims.title = "Claim Results"
    summaries = workbook.create_sheet("Article Summary")
    review = workbook.create_sheet("Review Queue")
    claim_headers = ["article_id", "published_at", "source_sentence", "claim_id", "verdict", "route_status", "reason_code", "claim_value", "calculated_value", "evidence_value", "kosis_table_id", "evidence_key", "source_url"]
    _write_sheet(claims, claim_headers, [[getattr(row, field) for field in claim_headers] for row in result.claim_rows])
    summary_headers = ["article_id", "published_at", "claim_count", "match_count", "mismatch_count", "review_count", "article_status", "source_url"]
    _write_sheet(summaries, summary_headers, [[getattr(row, field) for field in summary_headers] for row in result.article_rows])
    review_rows = [row for row in result.claim_rows if row.route_status != "AUTO"]
    _write_sheet(review, claim_headers, [[getattr(row, field) for field in claim_headers] for row in review_rows])
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _read_rows(suffix: str, content: bytes) -> list[dict[str, Any]]:
    if suffix == "csv":
        text = _decode_csv(content)
        try:
            dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\\t|")
        except csv.Error:
            dialect = csv.excel
        return [dict(row) for row in csv.DictReader(io.StringIO(text), dialect=dialect)]
    if suffix == "json":
        payload = json.loads(content.decode("utf-8"))
        if not isinstance(payload, list) or not all(isinstance(row, dict) for row in payload):
            raise ValueError("BATCH_FILE_INVALID")
        return payload
    if suffix == "xlsx":
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(value).strip() if value is not None else "" for value in rows[0]]
        return [dict(zip(headers, values, strict=False)) for values in rows[1:] if any(value is not None for value in values)]
    raise ValueError("BATCH_UNSUPPORTED_FILE")


_COLUMN_ALIASES = {
    "article_id": ("article_id", "articleid", "id", "기사id", "기사_id"),
    "published_at": ("published_at", "publisheddate", "article_date", "date", "기사일", "기사날짜", "발행일"),
    "body": ("body", "content", "sentence", "text", "본문", "문장", "기사본문", "내용", "기사내용", "뉴스본문", "claim", "claimtext"),
    "title": ("title", "headline", "제목"),
    "source_url": ("source_url", "url", "link", "원문url", "기사url"),
}


def _decode_csv(content: bytes) -> str:
    for encoding in ("utf-8-sig", "cp949"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError("BATCH_FILE_ENCODING")

def _canonicalize_row(row: dict[str, Any]) -> dict[str, Any]:
    normalized = {_normalize_header(key): value for key, value in row.items()}
    result = dict(row)
    for canonical, aliases in _COLUMN_ALIASES.items():
        for alias in aliases:
            if alias in normalized:
                result[canonical] = normalized[alias]
                break
    return result


def _normalize_header(value: object) -> str:
    return str(value).strip().replace(" ", "").replace("_", "").casefold()

def _numeric_claims(body: str) -> list[str]:
    sentences = [sentence.strip() for sentence in _SENTENCES.split(body) if sentence.strip()]
    return [claim for sentence in sentences if _NUMERIC.search(sentence) for claim in split_complex_claim(sentence)]


def _claim_row(article: BatchArticle, sentence: str, verdict: VerdictSchema) -> BatchClaimResult:
    cell = verdict.evidence_cells[0] if verdict.evidence_cells else None
    return BatchClaimResult(
        article_id=article.article_id,
        published_at=article.published_at,
        source_sentence=sentence,
        claim_id=verdict.claim_id,
        verdict=verdict.verdict,
        route_status=verdict.route_status,
        reason_code=verdict.reason_code,
        claim_value=verdict.claim_value,
        calculated_value=verdict.calculated_value,
        evidence_value=verdict.evidence_values[0] if verdict.evidence_values else None,
        kosis_table_id=cell.tbl_id if cell else None,
        evidence_key=cell.canonical_key if cell else None,
        source_url=article.source_url,
    )


def _hold_row(article: BatchArticle, sentence: str) -> BatchClaimResult:
    return BatchClaimResult(article.article_id, article.published_at, sentence, "batch_error", "UNDETERMINED", "HOLD", "BATCH_VERIFIER_ERROR", None, None, None, None, None, article.source_url)


def _summaries(articles: list[BatchArticle], rows: list[BatchClaimResult]) -> list[BatchArticleSummary]:
    grouped: dict[str, list[BatchClaimResult]] = defaultdict(list)
    for row in rows:
        grouped[row.article_id].append(row)
    summaries: list[BatchArticleSummary] = []
    for article in articles:
        article_rows = grouped[article.article_id]
        counts = Counter(row.verdict for row in article_rows)
        review_count = sum(row.route_status != "AUTO" for row in article_rows)
        status = "MISMATCH_FOUND" if counts["MISMATCH"] else "NEEDS_REVIEW" if review_count or not article_rows else "ALL_MATCH"
        summaries.append(BatchArticleSummary(article.article_id, article.published_at, len(article_rows), counts["MATCH"], counts["MISMATCH"], review_count, status, article.source_url))
    return summaries


def _write_sheet(sheet: Any, headers: list[str], rows: list[list[Any]]) -> None:
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    sheet.freeze_panes = "A2"
    for row in rows:
        sheet.append(row)
    for column in sheet.columns:
        sheet.column_dimensions[column[0].column_letter].width = min(max(len(str(cell.value or "")) for cell in column) + 2, 60)


def _required_text(row: dict[str, Any], key: str) -> str:
    value = _optional_text(row.get(key))
    if value is None:
        raise ValueError(key)
    return value


def _optional_text(value: Any) -> str | None:
    if value is None:
        return None
    normalized = str(value).strip()
    return normalized or None

