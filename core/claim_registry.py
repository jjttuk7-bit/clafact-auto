"""Build reproducible Claim Registry records from explicit source rows."""

from collections.abc import Iterable, Mapping
from datetime import date
import csv
import json
from pathlib import Path
from typing import Any

from schemas.claim import ClaimSchema
from schemas.claim_registry import ClaimRegistryRecord

_VALID_PARSE_STATUSES = {"AUTO_OK", "HOLD", "HUMAN_REVIEW"}


def build_registry_from_source(
    source_path: Path,
    *,
    source_ref: str,
    output_dir: Path,
    expected_count: int | None = None,
    source_sheet: str | None = None,
    header_row: int = 1,
    date_source_path: Path | None = None,
    date_sheet: str | None = None,
    date_header_row: int = 1,
) -> tuple[Path, Path]:
    """Create registry artifacts from a versioned CSV/XLSX source."""
    rows = load_registry_source_rows(
        source_path,
        source_sheet=source_sheet,
        header_row=header_row,
        date_source_path=date_source_path,
        date_sheet=date_sheet,
        date_header_row=date_header_row,
    )
    records, report = build_registry_records(
        rows, source_ref=source_ref, expected_count=expected_count
    )
    return write_registry_artifacts(records, report, output_dir=output_dir)


def load_registry_source_rows(
    source_path: Path,
    *,
    source_sheet: str | None = None,
    header_row: int = 1,
    date_source_path: Path | None = None,
    date_sheet: str | None = None,
    date_header_row: int = 1,
) -> list[dict[str, Any]]:
    """Read CSV/XLSX source rows and attach article publication dates when available."""
    rows = _read_tabular_rows(source_path, source_sheet, header_row)
    if date_source_path is None:
        return rows

    date_rows = _read_tabular_rows(date_source_path, date_sheet, date_header_row)
    dates_by_article = {
        _required_text(row, "article_id"): _optional_text(
            row, "article_published_at", "date", "작성일"
        )
        for row in date_rows
        if _optional_text(row, "article_id") is not None
    }
    return [
        {
            **row,
            "article_published_at": dates_by_article.get(
                _required_text(row, "article_id")
            ),
        }
        for row in rows
    ]


def build_registry_records(
    rows: Iterable[Mapping[str, Any]],
    *,
    source_ref: str,
    expected_count: int | None = None,
) -> tuple[list[ClaimRegistryRecord], dict[str, int | bool | None]]:
    """Convert source rows to a lossless, reviewable Claim Registry."""
    records: list[ClaimRegistryRecord] = []
    seen_keys: set[tuple[str, str]] = set()

    for row in rows:
        article_id = _required_text(row, "article_id")
        sentence_id = _required_text(row, "sentence_id")
        key = (article_id, sentence_id)
        if key in seen_keys:
            raise ValueError(f"Duplicate source key: {article_id}:{sentence_id}")
        seen_keys.add(key)

        source_sentence = _required_text(row, "sentence")
        claim = ClaimSchema(
            claim_id=_optional_text(row, "claim_id") or f"registry:{source_ref}:{article_id}:{sentence_id}",
            source_sentence=source_sentence,
            indicator=_optional_text(row, "claim_indicator", "indicator"),
            value=_optional_number(row, "claim_value", "value"),
            unit=_optional_text(row, "claim_unit", "unit"),
            time=_optional_text(row, "claim_time", "time"),
            frequency=_optional_text(row, "claim_frequency", "frequency"),
            region=_optional_text(row, "claim_region", "region"),
            population=_optional_text(row, "claim_population", "population"),
            dimension=_optional_dimension(row),
            comparison=_optional_mapping(row, "comparison"),
            calculation=_optional_calculation(row),
            condition=_optional_mapping(row, "condition"),
            source_hint=_optional_text(row, "source_hint"),
            parse_status=_parse_status(row),
        )
        records.append(
            ClaimRegistryRecord(
                article_id=article_id,
                sentence_id=sentence_id,
                article_published_at=_optional_date(row, "article_published_at", "article_date", "작성일"),
                source_ref=source_ref,
                source_metadata={
                    str(name): _optional_value(value) for name, value in row.items()
                },
                claim=claim,
            )
        )

    actual_count = len(records)
    return records, {
        "actual_count": actual_count,
        "expected_count": expected_count,
        "count_matches": expected_count is None or actual_count == expected_count,
        "duplicate_count": 0,
    }


def write_registry_artifacts(
    records: Iterable[ClaimRegistryRecord],
    report: Mapping[str, int | bool | None],
    *,
    output_dir: Path,
) -> tuple[Path, Path]:
    """Persist reproducible registry records and a separate validation report."""
    output_dir.mkdir(parents=True, exist_ok=True)
    jsonl_path = output_dir / "claim_registry.jsonl"
    report_path = output_dir / "validation_report.json"
    serialized_records = [
        json.dumps(record.model_dump(mode="json"), ensure_ascii=False, sort_keys=True)
        for record in records
    ]
    jsonl_path.write_text(
        "\n".join(serialized_records) + ("\n" if serialized_records else ""),
        encoding="utf-8",
    )
    report_path.write_text(
        json.dumps(dict(report), ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    return jsonl_path, report_path


def _read_tabular_rows(
    path: Path, sheet_name: str | None, header_row: int
) -> list[dict[str, Any]]:
    if path.suffix.lower() == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            return [dict(row) for row in csv.DictReader(handle)]
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError(f"Unsupported registry source format: {path.suffix}")
    if sheet_name is None:
        raise ValueError("An Excel source requires an explicit sheet name.")

    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name]
    rows = worksheet.iter_rows(min_row=header_row, values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    return [
        {headers[index]: value for index, value in enumerate(values) if headers[index]}
        for values in rows
        if any(value is not None and str(value).strip() for value in values)
    ]


def _required_text(row: Mapping[str, Any], key: str) -> str:
    value = _optional_text(row, key)
    if value is None:
        raise ValueError(f"Required source field is missing: {key}")
    return value


def _optional_text(row: Mapping[str, Any], *keys: str) -> str | None:
    for key in keys:
        value = row.get(key)
        normalized = _optional_value(value)
        if normalized is not None:
            return normalized
    return None


def _optional_value(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _optional_number(row: Mapping[str, Any], *keys: str) -> float | None:
    value = _optional_text(row, *keys)
    if value is None:
        return None
    normalized = value.replace(",", "").replace("%", "")
    try:
        return float(normalized)
    except ValueError:
        return None


def _optional_dimension(row: Mapping[str, Any]) -> dict[str, str] | None:
    value = _optional_text(row, "claim_dimension", "dimension")
    return {"raw": value} if value is not None else None


def _optional_mapping(row: Mapping[str, Any], *keys: str) -> dict[str, str] | None:
    value = _optional_text(row, *keys)
    if value is None:
        return None
    try:
        decoded = json.loads(value)
    except json.JSONDecodeError:
        return {"raw": value}
    if not isinstance(decoded, dict):
        return {"raw": value}
    return {str(key): str(item) for key, item in decoded.items() if item is not None}


def _optional_calculation(row: Mapping[str, Any]) -> str | None:
    value = _optional_text(row, "calculation")
    if value is None:
        return None
    try:
        decoded = json.loads(value)
    except json.JSONDecodeError:
        return value
    if isinstance(decoded, dict) and decoded.get("type") is not None:
        return str(decoded["type"])
    return value


def _optional_date(row: Mapping[str, Any], *keys: str) -> date | None:
    value = _optional_text(row, *keys)
    if value is None:
        return None
    try:
        return date.fromisoformat(value[:10])
    except ValueError:
        return None


def _parse_status(row: Mapping[str, Any]) -> str:
    value = _optional_text(row, "parse_review_status", "parse_status")
    return value if value in _VALID_PARSE_STATUSES else "HUMAN_REVIEW"
