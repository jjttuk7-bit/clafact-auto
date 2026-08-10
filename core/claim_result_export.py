"""In-memory download artifacts for one validated claim verdict."""

from __future__ import annotations

import json
from io import BytesIO

from openpyxl import Workbook

from schemas.verdict import VerdictSchema


def export_verdict_json_bytes(verdict: VerdictSchema) -> bytes:
    """Return the canonical verdict artifact without any provider payloads."""
    return (
        json.dumps(verdict.model_dump(mode="json"), ensure_ascii=False, indent=2, sort_keys=True)
        + "\n"
    ).encode("utf-8")


def export_verdict_xlsx_bytes(verdict: VerdictSchema) -> bytes:
    """Return a reviewer-ready workbook for one validated verdict."""
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    evidence = workbook.create_sheet("Evidence Cells")
    trace = workbook.create_sheet("Execution Trace")

    summary.append(["field", "value"])
    for field, value in verdict.model_dump(
        mode="json", exclude={"evidence_cells", "execution_trace"}
    ).items():
        summary.append(
            [
                field,
                json.dumps(value, ensure_ascii=False)
                if isinstance(value, (dict, list))
                else value,
            ]
        )

    evidence.append(
        [
            "index",
            "value",
            "org_id",
            "tbl_id",
            "itm_id",
            "prd_se",
            "prd_de",
            "unit",
            "canonical_key",
            "status",
            "dimension_codes",
        ]
    )
    for index, cell in enumerate(verdict.evidence_cells):
        evidence.append(
            [
                index,
                verdict.evidence_values[index]
                if index < len(verdict.evidence_values)
                else None,
                cell.org_id,
                cell.tbl_id,
                cell.itm_id,
                cell.prd_se,
                cell.prd_de,
                cell.unit,
                cell.canonical_key,
                cell.status,
                json.dumps(cell.dimension_codes, ensure_ascii=False, sort_keys=True),
            ]
        )

    trace.append(["stage", "status", "reason_code", "output_ref"])
    if verdict.execution_trace:
        for event in verdict.execution_trace.events:
            trace.append([event.stage, event.status, event.reason_code, event.output_ref])

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        for column in sheet.columns:
            sheet.column_dimensions[column[0].column_letter].width = min(
                max(len(str(cell.value or "")) for cell in column) + 2,
                60,
            )
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()