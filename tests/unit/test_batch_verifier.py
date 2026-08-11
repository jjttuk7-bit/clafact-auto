from datetime import date
from io import BytesIO

from openpyxl import load_workbook

import pytest

from core.batch_verifier import (
    BatchArticle,
    BatchClaimResult,
    export_batch_xlsx,
    load_articles,
    verify_articles,
)
from schemas.verdict import VerdictSchema


def _match_verdict(claim_id: str = "claim-1") -> VerdictSchema:
    return VerdictSchema(
        claim_id=claim_id,
        claim_value=28589000,
        evidence_values=[28589.0],
        calculated_value=28589000,
        verdict="MATCH",
        route_status="AUTO",
        reason_code="WITHIN_TOLERANCE",
        explanation="Claim matches the official calculation.",
        dataset_version="test",
        semantic_standard_version="test",
        kosis_catalog_version="test",
        matching_version="test",
        calculation_version="test",
    )


def test_load_articles_requires_required_columns() -> None:
    with pytest.raises(ValueError, match="BATCH_REQUIRED_COLUMNS"):
        load_articles("articles.csv", b"article_id,published_at\nA1,2025-04-09\n")


def test_load_articles_reads_csv_without_persisting_upload() -> None:
    articles = load_articles(
        "articles.csv",
        "article_id,published_at,title,body,source_url\nA1,2025-04-09,employment,2025년 3월 취업자 수는 2858만9000명이었다.,https://example.test/a1\n".encode(),
    )

    assert articles == [
        BatchArticle(
            article_id="A1",
            published_at=date(2025, 4, 9),
            body="2025년 3월 취업자 수는 2858만9000명이었다.",
            title="employment",
            source_url="https://example.test/a1",
        )
    ]


def test_verify_articles_creates_one_claim_result_per_numeric_sentence() -> None:
    article = BatchArticle("A1", date(2025, 4, 9), "배경이다. 2025년 3월 취업자 수는 2858만9000명이었다.")
    result = verify_articles([article], lambda sentence, _: _match_verdict())

    assert len(result.claim_rows) == 1
    assert result.claim_rows[0].article_id == "A1"
    assert result.claim_rows[0].source_sentence == "2025년 3월 취업자 수는 2858만9000명이었다."
    assert result.article_rows[0].article_status == "ALL_MATCH"


def test_verify_articles_splits_two_numeric_claims_in_one_sentence() -> None:
    article = BatchArticle("A1", date(2025, 4, 9), "2023년 고용률은 60%였고 2024년 고용률은 61%였다.")

    result = verify_articles([article], lambda sentence, _: _match_verdict(sentence))

    assert [row.source_sentence for row in result.claim_rows] == [
        "2023년 고용률은 60%",
        "2024년 고용률은 61%였다.",
    ]


def test_verify_articles_converts_verifier_failure_to_hold() -> None:
    article = BatchArticle("A1", date(2025, 4, 9), "2025년 3월 취업자 수는 2858만9000명이었다.")
    result = verify_articles([article], lambda sentence, _: (_ for _ in ()).throw(RuntimeError("failure")))

    assert result.claim_rows[0].route_status == "HOLD"
    assert result.claim_rows[0].reason_code == "BATCH_VERIFIER_ERROR"


def test_export_batch_xlsx_has_claim_summary_and_review_sheets() -> None:
    result = verify_articles(
        [BatchArticle("A1", date(2025, 4, 9), "2025년 3월 취업자 수는 2858만9000명이었다.")],
        lambda sentence, _: _match_verdict(),
    )

    payload = export_batch_xlsx(result)

    assert payload[:2] == b"PK"
    assert load_workbook(BytesIO(payload)).sheetnames == ["Claim Results", "Article Summary", "Review Queue"]


def test_load_articles_accepts_sentence_level_crawler_columns_with_default_date() -> None:
    articles = load_articles(
        "crawler.csv",
        "article_id,sentence_id,sentence,source_type\nA00006,1,2024년 수출은 6838억 달러였다.,KOSIS\n".encode(),
        default_published_at=date(2025, 4, 9),
    )

    assert articles[0].article_id == "A00006"
    assert articles[0].body == "2024년 수출은 6838억 달러였다."
    assert articles[0].published_at == date(2025, 4, 9)


def test_load_articles_requires_date_when_sentence_file_has_no_article_date() -> None:
    with pytest.raises(ValueError, match="BATCH_ARTICLE_DATE_REQUIRED"):
        load_articles("crawler.csv", b"article_id,sentence\nA00006,2024 data\n")


def test_load_articles_accepts_cp949_semicolon_file_without_article_id() -> None:
    articles = load_articles(
        "crawler.csv",
        "제목;내용\n고용;2025년 3월 취업자 수는 2858만9000명이었다.\n".encode("cp949"),
        default_published_at=date(2025, 4, 9),
    )

    assert articles[0].article_id == "row-000001"
    assert articles[0].body == "2025년 3월 취업자 수는 2858만9000명이었다."


def test_load_articles_accepts_article_body_crawler_headers() -> None:
    articles = load_articles(
        "news.csv",
        "기사제목,작성일,URL,기사 본문 전체\n물가 기사,2025-11-04,https://example.test/news,10월 소비자물가는 2.4% 상승했다.\n".encode(),
    )

    assert articles[0] == BatchArticle(
        article_id="row-000001",
        published_at=date(2025, 11, 4),
        body="10월 소비자물가는 2.4% 상승했다.",
        title="물가 기사",
        source_url="https://example.test/news",
    )


def test_extract_batch_claim_sentences_splits_statistical_claims_after_cleanup() -> None:
    from core.batch_verifier import extract_batch_claim_sentences

    body = '''경제 기사 제목 기자 입력 2025.11.04. 08:00 업데이트 2025.11.04. 11:11 5 지난달 소비자 물가가 2.4% 상승했고, 9월(2.1%)에 이어 2개월 연속 2%대였다.
이같은 물가 상승률은 지난해 7월(2.6%) 이후 15개월만에 가장 높다.
국가데이터처가 4일 발표한 자료에 따르면 물가는 1년 전보다 2.4% 올랐다.
소비자물가는 지난 8월 1.7%를 기록했다가 9월 2.1%로 올랐다.
가공식품 물가는 3.5% 오르며 전체 물가를 0.30%포인트 끌어올렸다.
9월 가공식품 물가 상승률은 4.2%였고 빵(6.6%), 커피(14.7%)도 올랐다.
외식 물가는 3.0% 상승했고 개인서비스 물가는 3.4% 올랐다.
축산물과 수산물 물가는 5.3%, 5.9% 올랐다.
농산물 물가는 1.1% 상승했다.
배추(-34.5%), 무(-40.5%), 쌀(21.3%), 사과(21.6%), 달걀(6.9%)의 상승률도 달랐다.
생활물가지수는 2.5% 올랐다.
신선식품지수는 0.8% 하락했다.
석유류 물가는 4.8% 상승했고 지난 2월(6.3%) 이후 가장 높았다.
먹거리 물가는 오름세였다. 관련 기사 광고 문구 500억원 할인 댓글 2025.11.05 11:03'''

    claims = extract_batch_claim_sentences(body)

    assert len(claims) == 21
    assert claims[0].startswith("지난달 소비자 물가")
    assert all("관련 기사" not in claim for claim in claims)
    assert all("업데이트" not in claim for claim in claims)


