import json
from io import BytesIO

from openpyxl import load_workbook

from core.claim_result_export import export_verdict_json_bytes, export_verdict_xlsx_bytes
from schemas.evidence import EvidenceCellSchema
from schemas.pipeline_trace import PipelineTraceSchema
from schemas.verdict import VerdictSchema


def _verdict() -> VerdictSchema:
    trace = PipelineTraceSchema(
        claim_id="claim-1",
        preprocess_version="1.0",
        claim_schema_version="1.0",
    ).pass_stage("CLAIM_PARSE")
    return VerdictSchema(
        claim_id="claim-1",
        claim_value=-34.5,
        evidence_values=[136.62, 208.57],
        calculated_value=-34.49681162199741,
        verdict="MATCH",
        route_status="AUTO",
        reason_code="WITHIN_TOLERANCE",
        explanation="Claim matches the official calculation.",
        evidence_cells=[
            EvidenceCellSchema(
                org_id="101",
                tbl_id="DT_1J22112",
                itm_id="T",
                dimension_codes={"C1": "T10", "C2": "A02A01701"},
                prd_se="M",
                prd_de="202510",
                unit="2020=100",
                canonical_key="DT_1J22112|T|202510|C1:T10|C2:A02A01701",
                status="CONFIRMED",
            )
        ],
        execution_trace=trace,
        dataset_version="goldset-v3-final",
        semantic_standard_version="1.0",
        kosis_catalog_version="1.0",
        matching_version="1.0",
        calculation_version="1.0",
    )


def test_export_verdict_json_preserves_verdict_and_trace() -> None:
    payload = json.loads(export_verdict_json_bytes(_verdict()).decode("utf-8"))

    assert payload["verdict"] == "MATCH"
    assert payload["execution_trace"]["claim_id"] == "claim-1"


def test_export_verdict_xlsx_has_summary_evidence_and_trace_sheets() -> None:
    workbook = load_workbook(BytesIO(export_verdict_xlsx_bytes(_verdict())))

    assert workbook.sheetnames == ["Summary", "Evidence Cells", "Execution Trace"]
    assert workbook["Evidence Cells"].max_row == 2
    assert workbook["Execution Trace"]["A2"].value == "CLAIM_PARSE"