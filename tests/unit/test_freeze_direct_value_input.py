import csv
import hashlib
import json
from pathlib import Path

import pytest
from openpyxl import Workbook

from tools.freeze_direct_value_input import freeze_sheet


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest().upper()


def _sample_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "08_직접값"
    sheet.append(["직접값 세부분류 실행원장 — 2건"])
    sheet.append(["설명"])
    sheet.append(["완료 조건"])
    sheet.append(["실행순서", "Claim번호", "원문", "기사값"])
    sheet.append([1, "A001_1", "출생아 수는 23만 명이다.", 230000])
    sheet.append([2, "A002_1", "취업자는 10만 명이다.", 100000])
    workbook.save(path)


def test_freeze_sheet_preserves_headers_rows_and_records_hashes(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    output_dir = tmp_path / "frozen"
    _sample_workbook(source)

    result = freeze_sheet(
        source_workbook=source,
        sheet_name="08_직접값",
        header_row=4,
        expected_rows=2,
        output_dir=output_dir,
        frozen_at="2026-08-25T17:00:00+09:00",
        code_revision="abc123",
    )

    csv_path = Path(result["outputs"]["csv"]["path"])
    jsonl_path = Path(result["outputs"]["jsonl"]["path"])
    manifest_path = Path(result["manifest_path"])

    assert csv_path.read_bytes().startswith(b"\xef\xbb\xbf")
    with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    assert list(rows[0]) == ["실행순서", "Claim번호", "원문", "기사값"]
    assert [row["Claim번호"] for row in rows] == ["A001_1", "A002_1"]

    jsonl_rows = [json.loads(line) for line in jsonl_path.read_text(encoding="utf-8").splitlines()]
    assert jsonl_rows[0]["기사값"] == 230000
    assert jsonl_rows[1]["원문"] == "취업자는 10만 명이다."

    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    assert manifest["row_count"] == 2
    assert manifest["column_count"] == 4
    assert manifest["formula_count"] == 0
    assert manifest["source_workbook"]["sha256"] == _sha256(source)
    assert manifest["outputs"]["csv"]["sha256"] == _sha256(csv_path)
    assert manifest["outputs"]["jsonl"]["sha256"] == _sha256(jsonl_path)
    assert manifest["code_revision"] == "abc123"


def test_freeze_sheet_fails_closed_when_row_count_differs(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    _sample_workbook(source)

    with pytest.raises(ValueError, match="expected 381 rows, found 2"):
        freeze_sheet(
            source_workbook=source,
            sheet_name="08_직접값",
            header_row=4,
            expected_rows=381,
            output_dir=tmp_path / "frozen",
        )


def test_freeze_sheet_rejects_formula_cells(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    _sample_workbook(source)
    from openpyxl import load_workbook

    workbook = load_workbook(source)
    workbook["08_직접값"]["D5"] = "=100+200"
    workbook.save(source)

    with pytest.raises(ValueError, match="formula cells are not allowed"):
        freeze_sheet(
            source_workbook=source,
            sheet_name="08_직접값",
            header_row=4,
            expected_rows=2,
            output_dir=tmp_path / "frozen",
        )
