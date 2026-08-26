"""Freeze one workbook sheet into deterministic CSV and JSONL inputs."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import subprocess
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


CSV_NAME = "CLAFACT_8번_직접값_381건_동결입력.csv"
JSONL_NAME = "CLAFACT_8번_직접값_381건_동결입력.jsonl"
MANIFEST_NAME = "CLAFACT_8번_직접값_381건_동결_manifest.json"


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest().upper()


def _json_value(value: Any) -> Any:
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, float) and not math.isfinite(value):
        raise ValueError("non-finite numeric cell is not allowed")
    return value


def _file_record(path: Path) -> dict[str, Any]:
    return {
        "path": str(path.resolve()),
        "bytes": path.stat().st_size,
        "sha256": _sha256(path),
    }


def freeze_sheet(
    *,
    source_workbook: Path,
    sheet_name: str,
    header_row: int,
    expected_rows: int,
    output_dir: Path,
    frozen_at: str | None = None,
    code_revision: str | None = None,
) -> dict[str, Any]:
    source_workbook = Path(source_workbook).resolve()
    output_dir = Path(output_dir).resolve()

    formula_book = load_workbook(source_workbook, read_only=True, data_only=False)
    try:
        if sheet_name not in formula_book.sheetnames:
            raise ValueError(f"sheet not found: {sheet_name}")
        formula_sheet = formula_book[sheet_name]
        formula_cells = [
            cell.coordinate
            for row in formula_sheet.iter_rows(min_row=header_row + 1)
            for cell in row
            if cell.data_type == "f"
        ]
    finally:
        formula_book.close()
    if formula_cells:
        raise ValueError(f"formula cells are not allowed: {', '.join(formula_cells[:10])}")

    workbook = load_workbook(source_workbook, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name]
        header_cells = next(sheet.iter_rows(min_row=header_row, max_row=header_row))
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in header_cells]
        while headers and not headers[-1]:
            headers.pop()
        if not headers or any(not header for header in headers):
            raise ValueError("headers must be non-empty")
        if len(set(headers)) != len(headers):
            raise ValueError("headers must be unique")

        records: list[dict[str, Any]] = []
        for cells in sheet.iter_rows(min_row=header_row + 1, max_col=len(headers)):
            values = [_json_value(cell.value) for cell in cells]
            if not any(value is not None and str(value).strip() for value in values):
                continue
            records.append(dict(zip(headers, values, strict=True)))
    finally:
        workbook.close()

    if len(records) != expected_rows:
        raise ValueError(f"expected {expected_rows} rows, found {len(records)}")

    output_dir.mkdir(parents=True, exist_ok=True)
    csv_path = output_dir / CSV_NAME
    jsonl_path = output_dir / JSONL_NAME
    manifest_path = output_dir / MANIFEST_NAME

    with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers, extrasaction="raise")
        writer.writeheader()
        writer.writerows(records)

    with jsonl_path.open("w", encoding="utf-8", newline="\n") as handle:
        for record in records:
            handle.write(json.dumps(record, ensure_ascii=False, separators=(",", ":"), allow_nan=False))
            handle.write("\n")

    if frozen_at is None:
        frozen_at = datetime.now().astimezone().isoformat(timespec="seconds")
    if code_revision is None:
        code_revision = subprocess.run(
            ["git", "rev-parse", "HEAD"],
            cwd=Path(__file__).resolve().parents[1],
            check=True,
            capture_output=True,
            text=True,
        ).stdout.strip()

    manifest: dict[str, Any] = {
        "version": 1,
        "frozen_at": frozen_at,
        "code_revision": code_revision,
        "source_workbook": {
            "path": str(source_workbook),
            "bytes": source_workbook.stat().st_size,
            "last_modified": datetime.fromtimestamp(source_workbook.stat().st_mtime).astimezone().isoformat(),
            "sha256": _sha256(source_workbook),
        },
        "sheet_name": sheet_name,
        "header_row": header_row,
        "data_start_row": header_row + 1,
        "row_count": len(records),
        "column_count": len(headers),
        "headers": headers,
        "formula_count": 0,
        "outputs": {
            "csv": _file_record(csv_path),
            "jsonl": _file_record(jsonl_path),
        },
    }
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    result = dict(manifest)
    result["manifest_path"] = str(manifest_path)
    return result


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("source_workbook", type=Path)
    parser.add_argument("--sheet", default="08_직접값")
    parser.add_argument("--header-row", type=int, default=4)
    parser.add_argument("--expected-rows", type=int, default=381)
    parser.add_argument("--output-dir", type=Path, required=True)
    args = parser.parse_args()
    result = freeze_sheet(
        source_workbook=args.source_workbook,
        sheet_name=args.sheet,
        header_row=args.header_row,
        expected_rows=args.expected_rows,
        output_dir=args.output_dir,
    )
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
